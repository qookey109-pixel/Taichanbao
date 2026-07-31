#!/usr/bin/env python3
"""Import CSV, TSV, and XLSX files from incoming/ into the project SQLite database."""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable


PROJECT_DIR = Path(__file__).resolve().parent.parent
INCOMING_DIR = PROJECT_DIR / "incoming"
DATABASE_PATH = PROJECT_DIR / "db" / "taiwan_industry_report.sqlite"
SCHEMA_PATH = PROJECT_DIR / "db" / "schema.sql"
SUPPORTED_SUFFIXES = {".csv", ".tsv", ".xlsx"}


def quote_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def safe_identifier(value: object, fallback: str) -> str:
    value = re.sub(r"[^0-9A-Za-z_]+", "_", str(value or "").strip())
    value = re.sub(r"_+", "_", value).strip("_").lower()
    if not value:
        value = fallback
    if value[0].isdigit():
        value = f"field_{value}"
    return value[:60]


def unique_headers(headers: Iterable[object]) -> list[str]:
    used: set[str] = set()
    result: list[str] = []
    for position, header in enumerate(headers, start=1):
        base = safe_identifier(header, f"column_{position}")
        candidate = base
        counter = 2
        while candidate in used:
            candidate = f"{base}_{counter}"
            counter += 1
        used.add(candidate)
        result.append(candidate)
    return result


def digest(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def read_delimited(path: Path, delimiter: str) -> tuple[list[object], list[list[object]]]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp950", "big5"):
        try:
            with path.open("r", encoding=encoding, newline="") as source:
                rows = list(csv.reader(source, delimiter=delimiter))
            break
        except UnicodeDecodeError as error:
            last_error = error
    else:
        raise RuntimeError(f"Cannot read {path.name}; unsupported text encoding.") from last_error
    if not rows:
        return [], []
    return rows[0], rows[1:]


def read_xlsx(path: Path) -> list[tuple[str, list[object], list[list[object]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "Excel import requires openpyxl. Install it with: python3 -m pip install openpyxl"
        ) from error
    workbook = load_workbook(path, read_only=True, data_only=True)
    datasets = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if rows:
            datasets.append((sheet.title, list(rows[0]), [list(row) for row in rows[1:]]))
    return datasets


def ensure_table(connection: sqlite3.Connection, table: str, headers: list[str]) -> None:
    columns = [f"{quote_identifier('_import_id')} INTEGER NOT NULL", f"{quote_identifier('_source_row')} INTEGER NOT NULL"]
    columns.extend(f"{quote_identifier(header)} TEXT" for header in headers)
    connection.execute(
        f"CREATE TABLE IF NOT EXISTS {quote_identifier(table)} ({', '.join(columns)})"
    )
    existing_columns = {
        row[1] for row in connection.execute(f"PRAGMA table_info({quote_identifier(table)})")
    }
    for header in headers:
        if header not in existing_columns:
            connection.execute(
                f"ALTER TABLE {quote_identifier(table)} ADD COLUMN {quote_identifier(header)} TEXT"
            )


def import_dataset(
    connection: sqlite3.Connection,
    source_path: Path,
    source_hash: str,
    sheet_name: str | None,
    headers: list[object],
    rows: list[list[object]],
    force: bool,
) -> str:
    source_label = source_path.stem if sheet_name is None else f"{source_path.stem}_{sheet_name}"
    table = safe_identifier(source_label, "imported_data")
    existing_import = connection.execute(
        "SELECT id, target_table FROM import_log WHERE source_hash = ? AND sheet_name IS ?",
        (source_hash, sheet_name),
    ).fetchone()
    if existing_import and not force:
        return f"skip  {source_path.name}{' / ' + sheet_name if sheet_name else ''} (unchanged)"

    normalized_headers = unique_headers(headers)
    if not normalized_headers:
        return f"skip  {source_path.name}{' / ' + sheet_name if sheet_name else ''} (empty)"
    ensure_table(connection, table, normalized_headers)

    if existing_import:
        previous_id, previous_table = existing_import
        connection.execute(
            f"DELETE FROM {quote_identifier(previous_table)} WHERE {quote_identifier('_import_id')} = ?",
            (previous_id,),
        )
        connection.execute("DELETE FROM import_log WHERE id = ?", (previous_id,))
    insert_log = connection.execute(
        """INSERT INTO import_log
           (source_path, source_hash, sheet_name, target_table, row_count, imported_at)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (str(source_path.relative_to(PROJECT_DIR)), source_hash, sheet_name, table, len(rows), datetime.now(timezone.utc).isoformat()),
    )
    import_id = insert_log.lastrowid

    insert_columns = ["_import_id", "_source_row", *normalized_headers]
    placeholders = ", ".join("?" for _ in insert_columns)
    column_names = ", ".join(quote_identifier(column) for column in insert_columns)
    values = []
    for index, row in enumerate(rows, start=2):
        padded = list(row[: len(normalized_headers)]) + [None] * max(0, len(normalized_headers) - len(row))
        values.append([import_id, index, *[None if value is None else str(value) for value in padded]])
    if values:
        connection.executemany(
            f"INSERT INTO {quote_identifier(table)} ({column_names}) VALUES ({placeholders})", values
        )
    return f"import {source_path.name}{' / ' + sheet_name if sheet_name else ''} → {table} ({len(values)} rows)"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--force", action="store_true", help="Import unchanged files again")
    args = parser.parse_args()
    INCOMING_DIR.mkdir(exist_ok=True)
    DATABASE_PATH.parent.mkdir(exist_ok=True)
    connection = sqlite3.connect(DATABASE_PATH)
    try:
        connection.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
        files = sorted(path for path in INCOMING_DIR.iterdir() if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES)
        if not files:
            print(f"Database ready: {DATABASE_PATH}")
            print("No source files found in incoming/.")
            return 0
        for path in files:
            file_hash = digest(path)
            if path.suffix.lower() == ".xlsx":
                datasets = read_xlsx(path)
            else:
                delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
                header, rows = read_delimited(path, delimiter)
                datasets = [(None, header, rows)]
            for sheet_name, headers, rows in datasets:
                print(import_dataset(connection, path, file_hash, sheet_name, headers, rows, args.force))
        connection.commit()
    finally:
        connection.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
